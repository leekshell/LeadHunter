"""
LeadHunter - Prospecção Automática de Leads
===========================================

Versão 2.0 (build comercial)
----------------------------

Aplicação desktop para prospectar empresas em um nicho + cidade,
enriquecer os contatos a partir do site público da empresa e exportar
os resultados em Excel (com foco em WhatsApp e redes sociais).

Este software foi desenvolvido para uso profissional e deve ser
empregado de forma responsável, respeitando os Termos de Uso dos
serviços consultados (Google Maps, sites das empresas) e as normas
de proteção de dados aplicáveis (no Brasil, a LGPD - Lei 13.709/2018).

Licença: MIT (ver LICENSE).
Mais informações: README.md e TERMOS_DE_USO_E_PRIVACIDADE.md.
"""

import asyncio
import io
import json
import logging
import logging.handlers
import os
import random
import re
import sqlite3
import sys
import threading
import time
import warnings
from collections import deque
from datetime import datetime
from pathlib import Path
from urllib.parse import quote_plus, urlparse, urlunparse

# --- CORREÇÃO DE NAVEGADORES PLAYWRIGHT EM EXECUTÁVEIS (.EXE) ---
if getattr(sys, "frozen", False):
    os.environ["PLAYWRIGHT_BROWSERS_PATH"] = os.path.join(
        os.environ.get("LOCALAPPDATA", ""), "ms-playwright"
    )
else:
    os.environ["PLAYWRIGHT_BROWSERS_PATH"] = "0"

# Silencia avisos de DeprecationWarning do Python 3.14
warnings.filterwarnings("ignore", category=DeprecationWarning)

import customtkinter as ctk
import httpx
import psutil
from bs4 import BeautifulSoup
from PIL import Image
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from playwright.async_api import async_playwright

ctk.set_appearance_mode("dark")
ctk.set_default_color_theme("blue")

# ---------------------------------------------------------------------------
# CONSTANTES DA APLICAÇÃO
# ---------------------------------------------------------------------------
APP_NAME = "LeadHunter"
APP_VERSION = "2.0.0"
APP_BUILD = "build comercial"
APP_STAGE = "Estável"

EMAIL_REGEX = re.compile(r"[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}")
PHONE_REGEX = re.compile(r"\D")
NOISE_EMAILS = (
    "sentry", "wixpress", "example", ".png", ".jpg", ".jpeg", ".gif",
    ".webp", "bootstrap", "schema.org", "placeholder", "no-reply",
    "noreply", "email.com", "yourdomain", "domain.com",
)
PLACEHOLDER_NAMES = {
    "", "n/d", "resultados", "pesquisar neste local",
    "pesquisar aqui", "google maps", "mapas do google",
}
DEFAULT_CONFIG = {
    "nicho": "",
    "cidades": "",
    "max_perfis_por_cidade": 60,
    "delay_entre_perfis": 1.2,
    "delay_entre_cidades": 8.0,
    "jitter": 0.6,
    "coletar_site": True,
    "coletar_redes": True,
    "modo_headless": True,
    "confirmar_uso_responsavel": True,
}


# ---------------------------------------------------------------------------
# UTILITÁRIOS
# ---------------------------------------------------------------------------
def _normalize_url(url: str) -> str:
    """Remove query string e fragmento para gerar um id estável de lead."""
    if not url:
        return url
    try:
        parsed = urlparse(url)
        return urlunparse((parsed.scheme, parsed.netloc, parsed.path, "", "", ""))
    except Exception:
        return url.split("?")[0]


def _extract_phone(raw: str) -> str:
    if not raw:
        return "N/D"
    digits = PHONE_REGEX.sub("", raw)
    if digits.startswith("0"):
        digits = digits[1:]
    return digits if digits else "N/D"


def _is_noise_email(email: str) -> bool:
    email = (email or "").lower()
    return any(n in email for n in NOISE_EMAILS)


def _is_placeholder_name(name: str) -> bool:
    return (name or "").strip().lower() in PLACEHOLDER_NAMES


def _safe_float(value, default: float) -> float:
    try:
        return float(value)
    except (TypeError, ValueError):
        return default


def _safe_int(value, default: int) -> int:
    try:
        return int(value)
    except (TypeError, ValueError):
        return default


class AppConfig:
    """Configuração persistente da aplicação (config.json)."""

    def __init__(self, path: Path):
        self.path = path
        self.data = dict(DEFAULT_CONFIG)
        self._load()

    def _load(self):
        if not self.path.exists():
            return
        try:
            saved = json.loads(self.path.read_text(encoding="utf-8"))
            for key, value in saved.items():
                if key in DEFAULT_CONFIG:
                    self.data[key] = value
        except Exception:
            pass

    def save(self):
        try:
            self.path.parent.mkdir(parents=True, exist_ok=True)
            self.path.write_text(
                json.dumps(self.data, indent=2, ensure_ascii=False),
                encoding="utf-8",
            )
        except Exception:
            pass

    def get(self, key, default=None):
        return self.data.get(key, default)

    def set(self, key, value):
        self.data[key] = value


# ---------------------------------------------------------------------------
# APLICAÇÃO PRINCIPAL
# ---------------------------------------------------------------------------
class LeadHunterApp(ctk.CTk):
    def __init__(self):
        super().__init__()

        self.title(f"{APP_NAME} - Prospecção Automática de Leads v{APP_VERSION}")
        self.geometry("1180x820")
        self.minsize(1000, 680)

        # ------------------ GERENCIAMENTO DE PASTAS LOCAIS (.EXE / .PY) ----
        if getattr(sys, "frozen", False):
            self.base_dir = Path(sys.executable).parent.resolve()
        else:
            self.base_dir = Path(__file__).parent.resolve() if "__file__" in globals() else Path.cwd()

        self.logs_dir = self.base_dir / "logs"
        self.relatorios_dir = self.base_dir / "relatorios"
        self.logs_dir.mkdir(parents=True, exist_ok=True)
        self.relatorios_dir.mkdir(parents=True, exist_ok=True)

        self.config = AppConfig(self.base_dir / "config.json")
        self.db_path = self.base_dir / "leadhunter_storage.db"

        # ------------------ CONTROLE DE EXECUÇÃO ---------------------------
        self.is_running = False
        self.ram_buffer = deque()
        self._stop_requested = False

        self._setup_logging()
        self.init_db()
        self.build_ui()
        self.start_ram_monitor()

        self.log_message(
            f"[SISTEMA] {APP_NAME} v{APP_VERSION} ({APP_BUILD}) | {datetime.now():%Y-%m-%d %H:%M:%S}"
        )
        self.log_message(f"[SISTEMA] Diretório local de execução: {self.base_dir}")
        self.log_message(f"[SISTEMA] Pasta de Logs pronta: {self.logs_dir}")
        self.log_message(f"[SISTEMA] Pasta de Relatórios pronta: {self.relatorios_dir}")
        self.log_message("[SISTEMA] Use o software apenas para contatos comerciais legítimos, respeitando a LGPD.")
        self.update_status("Pronto para iniciar.")

    # ------------------------------------------------------------------
    # LOG / THREAD-SAFE UI
    # ------------------------------------------------------------------
    def _setup_logging(self):
        log_file = self.logs_dir / "leadhunter.log"
        handler = logging.handlers.RotatingFileHandler(
            log_file, maxBytes=5 * 1024 * 1024, backupCount=5, encoding="utf-8"
        )
        handler.setFormatter(logging.Formatter("%(asctime)s [%(levelname)s] %(message)s"))
        self.logger = logging.getLogger(APP_NAME)
        self.logger.handlers.clear()
        self.logger.addHandler(handler)
        self.logger.setLevel(logging.INFO)

    def _ui(self, callback, *args, **kwargs):
        """Executa uma atualização de interface com segurança a partir de threads."""
        try:
            if threading.current_thread() is threading.main_thread():
                callback(*args, **kwargs)
            else:
                self.after(0, lambda: callback(*args, **kwargs))
        except Exception:
            pass

    # ------------------------------------------------------------------
    # BANCO DE DADOS
    # ------------------------------------------------------------------
    def init_db(self):
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        cursor.execute(
            """
            CREATE TABLE IF NOT EXISTS leads (
                id TEXT PRIMARY KEY,
                nome TEXT,
                telefone TEXT,
                email TEXT,
                site TEXT,
                whatsapp TEXT,
                instagram TEXT,
                facebook TEXT,
                linkedin TEXT,
                nicho TEXT,
                cidade TEXT,
                timestamp TEXT
            )
            """
        )
        cursor.execute(
            "CREATE INDEX IF NOT EXISTS idx_leads_nicho_cidade ON leads(nicho, cidade)"
        )
        cursor.execute("CREATE INDEX IF NOT EXISTS idx_leads_timestamp ON leads(timestamp)")
        cursor.execute(
            """
            CREATE TABLE IF NOT EXISTS app_settings (
                key TEXT PRIMARY KEY,
                value TEXT
            )
            """
        )
        conn.commit()
        conn.close()

    def is_lead_processed(self, lead_id: str) -> bool:
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        cursor.execute("SELECT 1 FROM leads WHERE id = ?", (lead_id,))
        exists = cursor.fetchone() is not None
        conn.close()
        return exists

    def flush_buffer_to_db(self):
        if not self.ram_buffer:
            return
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        while self.ram_buffer:
            item = self.ram_buffer.popleft()
            cursor.execute(
                """
                INSERT OR IGNORE INTO leads
                (id, nome, telefone, email, site, whatsapp, instagram, facebook,
                 linkedin, nicho, cidade, timestamp)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    item["id"], item["nome"], item["telefone"], item["email"],
                    item["site"], item["whatsapp"], item["instagram"], item["facebook"],
                    item["linkedin"], item["nicho"], item["cidade"], item["timestamp"],
                ),
            )
        conn.commit()
        conn.close()

    # ------------------------------------------------------------------
    # INTERFACE
    # ------------------------------------------------------------------
    def build_ui(self):
        self.grid_columnconfigure(0, weight=1)
        self.grid_columnconfigure(1, weight=1)
        self.grid_rowconfigure(0, weight=1)
        self.grid_rowconfigure(4, weight=2)

        # -------- COLUNA ESQUERDA: FILTROS --------
        self.frame_inputs = ctk.CTkFrame(self, corner_radius=12)
        self.frame_inputs.grid(row=0, column=0, padx=15, pady=15, sticky="nsew")

        ctk.CTkLabel(
            self.frame_inputs,
            text="Filtros de Prospecção",
            font=("Segoe UI", 16, "bold"),
        ).pack(anchor="w", padx=15, pady=(12, 5))

        self.var_nicho = ctk.StringVar(value=self.config.get("nicho", ""))
        self.entry_nicho = ctk.CTkEntry(
            self.frame_inputs, textvariable=self.var_nicho,
            placeholder_text="Nicho (Ex: Odontologia)",
        )
        self.entry_nicho.pack(fill="x", padx=15, pady=5)

        self.var_cidades = ctk.StringVar(value=self.config.get("cidades", ""))
        self.entry_cidade = ctk.CTkEntry(
            self.frame_inputs, textvariable=self.var_cidades,
            placeholder_text="Cidades (Ex: Campinas, Sumaré, Artur Nogueira)",
        )
        self.entry_cidade.pack(fill="x", padx=15, pady=5)

        ctk.CTkLabel(
            self.frame_inputs,
            text="Dicas de uso",
            font=("Segoe UI", 12, "bold"),
        ).pack(anchor="w", padx=15, pady=(12, 2))

        dicas = (
            "• Separe as cidades por vírgula.\n"
            "• Prefira lotes pequenos para reduzir bloqueios.\n"
            "• Respeite a pausa configurada entre perfis.\n"
            "• Use apenas para contatos comerciais legítimos."
        )
        self.lbl_dicas = ctk.CTkLabel(
            self.frame_inputs, text=dicas, justify="left", anchor="w", font=("Segoe UI", 11)
        )
        self.lbl_dicas.pack(anchor="w", padx=15, pady=(0, 12))

        # -------- COLUNA DIREITA: MINI PLAYER --------
        self.frame_player = ctk.CTkFrame(self, corner_radius=12)
        self.frame_player.grid(row=0, column=1, padx=15, pady=15, sticky="nsew")

        ctk.CTkLabel(
            self.frame_player,
            text="Monitor em Tempo Real",
            font=("Segoe UI", 14, "bold"),
        ).pack(anchor="w", padx=10, pady=8)

        self.lbl_mini_player = ctk.CTkLabel(
            self.frame_player, text="Navegador em Espera",
            width=420, height=240, fg_color="#1e293b", corner_radius=8,
        )
        self.lbl_mini_player.pack(expand=True, fill="both", padx=10, pady=10)

        ctk.CTkLabel(
            self.frame_player,
            text="Compliance: uso responsável e conforme a LGPD.",
            font=("Segoe UI", 10, "italic"),
            text_color="#94a3b8",
        ).pack(anchor="w", padx=10, pady=(0, 10))

        # -------- CONFIGURAÇÕES DE EXECUÇÃO --------
        self.frame_settings = ctk.CTkFrame(self, corner_radius=10)
        self.frame_settings.grid(row=1, column=0, columnspan=2, padx=15, pady=(0, 10), sticky="ew")
        for col in (0, 1, 2):
            self.frame_settings.grid_columnconfigure(col, weight=1)

        ctk.CTkLabel(
            self.frame_settings,
            text="Configurações de Execução",
            font=("Segoe UI", 14, "bold"),
        ).grid(row=0, column=0, columnspan=3, sticky="w", padx=15, pady=(10, 4))

        self.var_max_perfis = ctk.StringVar(value=str(self.config.get("max_perfis_por_cidade", 60)))
        self.var_delay_perfil = ctk.StringVar(value=str(self.config.get("delay_entre_perfis", 1.2)))
        self.var_delay_cidade = ctk.StringVar(value=str(self.config.get("delay_entre_cidades", 8.0)))

        self._build_setting_row(
            row=1, col=0, label="Máx. perfis/cidade",
            var=self.var_max_perfis, tooltip="Limita a quantidade de perfis processados por cidade."
        )
        self._build_setting_row(
            row=1, col=1, label="Pausa entre perfis (s)",
            var=self.var_delay_perfil, tooltip="Pausa mínima entre um perfil e outro (anti-bloqueio)."
        )
        self._build_setting_row(
            row=1, col=2, label="Pausa entre cidades (s)",
            var=self.var_delay_cidade, tooltip="Pausa entre uma cidade e a próxima."
        )

        self.var_coletar_site = ctk.BooleanVar(value=bool(self.config.get("coletar_site", True)))
        self.var_coletar_redes = ctk.BooleanVar(value=bool(self.config.get("coletar_redes", True)))
        self.var_headless = ctk.BooleanVar(value=bool(self.config.get("modo_headless", True)))
        self.var_confirmar = ctk.BooleanVar(value=bool(self.config.get("confirmar_uso_responsavel", True)))

        self.chk_coletar_site = ctk.CTkCheckBox(
            self.frame_settings, text="Coletar dados do site",
            variable=self.var_coletar_site,
        )
        self.chk_coletar_site.grid(row=2, column=0, padx=15, pady=4, sticky="w")

        self.chk_coletar_redes = ctk.CTkCheckBox(
            self.frame_settings, text="Coletar redes sociais",
            variable=self.var_coletar_redes,
        )
        self.chk_coletar_redes.grid(row=2, column=1, padx=15, pady=4, sticky="w")

        self.chk_headless = ctk.CTkCheckBox(
            self.frame_settings, text="Navegador invisível",
            variable=self.var_headless,
        )
        self.chk_headless.grid(row=2, column=2, padx=15, pady=4, sticky="w")

        self.chk_confirmar = ctk.CTkCheckBox(
            self.frame_settings,
            text="Confirmo uso responsável e conforme a LGPD",
            variable=self.var_confirmar,
        )
        self.chk_confirmar.grid(row=3, column=0, columnspan=2, padx=15, pady=4, sticky="w")

        self.btn_legal = ctk.CTkButton(
            self.frame_settings, text="Termos & Privacidade",
            fg_color="#334155", hover_color="#475569", width=120,
            command=self.open_legal_info,
        )
        self.btn_legal.grid(row=3, column=2, padx=15, pady=4, sticky="e")

        # -------- STATUS DE RAM --------
        self.frame_ram_bar = ctk.CTkFrame(self, corner_radius=8)
        self.frame_ram_bar.grid(row=2, column=0, columnspan=2, padx=15, pady=(0, 10), sticky="ew")

        self.lbl_ram_status = ctk.CTkLabel(
            self.frame_ram_bar,
            text="Uso de Memória RAM do Sistema: 0%",
            font=("Segoe UI", 11, "bold"),
        )
        self.lbl_ram_status.pack(anchor="w", padx=15, pady=(5, 2))

        self.progress_ram = ctk.CTkProgressBar(self.frame_ram_bar, height=12)
        self.progress_ram.set(0.0)
        self.progress_ram.pack(fill="x", padx=15, pady=(0, 8))

        # -------- CONTROLES --------
        self.frame_controls = ctk.CTkFrame(self, corner_radius=8)
        self.frame_controls.grid(row=3, column=0, columnspan=2, padx=15, pady=5, sticky="ew")

        self.btn_start = ctk.CTkButton(
            self.frame_controls,
            text="Iniciar Varredura",
            command=self.toggle_scraping,
            fg_color="#10b981", hover_color="#059669",
            font=("Segoe UI", 12, "bold"),
        )
        self.btn_start.pack(side="left", padx=15, pady=10)

        self.lbl_progress_status = ctk.CTkLabel(
            self.frame_controls,
            text="Status: Aguardando início...",
            font=("Segoe UI", 12, "bold"),
        )
        self.lbl_progress_status.pack(side="left", padx=15)

        self.progress_search = ctk.CTkProgressBar(self.frame_controls, width=300)
        self.progress_search.set(0.0)
        self.progress_search.pack(side="right", padx=15, pady=10)

        # -------- LOG --------
        self.textbox_log = ctk.CTkTextbox(
            self, corner_radius=10, font=("Consolas", 11), wrap="word"
        )
        self.textbox_log.grid(row=4, column=0, columnspan=2, padx=15, pady=(0, 15), sticky="nsew")
        self.grid_rowconfigure(4, weight=2)

    def _build_setting_row(self, row, col, label, var, tooltip):
        frame = ctk.CTkFrame(self.frame_settings, fg_color="transparent")
        frame.grid(row=row, column=col, sticky="nsew", padx=15, pady=4)
        frame.grid_columnconfigure(0, weight=1)
        frame.grid_columnconfigure(1, weight=0)

        ctk.CTkLabel(frame, text=label).grid(row=0, column=0, padx=4, pady=4, sticky="w")
        entry = ctk.CTkEntry(frame, width=130, textvariable=var)
        entry.grid(row=0, column=1, padx=4, pady=4, sticky="e")
        ctk.CTkLabel(frame, text=tooltip, text_color="#94a3b8", font=("Segoe UI", 10)).grid(
            row=1, column=0, columnspan=2, padx=4, pady=2, sticky="w"
        )

    def open_legal_info(self):
        text = (
            "USO RESPONSÁVEL E LGPD\n\n"
            "1. Esta ferramenta deve ser usada apenas para contatos comerciais "
            "legítimos de empresas.\n"
            "2. Os dados coletados são de acesso público e devem ser tratados com "
            "finalidade legítima (ex.: B2B) e boa-fé.\n"
            "3. Você é o controlador dos dados e é responsável por obter a base legal "
            "adequada, manter registros e atender a pedidos de exclusão.\n"
            "4. Não use esta ferramenta para spam, envio não solicitado em massa "
            "ou captação indevida de dados pessoais.\n"
            "5. Respeite os Termos de Uso dos sites consultados e as taxas de acesso.\n\n"
            "Consulte TERMOS_DE_USO_E_PRIVACIDADE.md para detalhes."
        )
        self.log_message("[LEGAL] Exibindo informações de uso responsável.")
        self._show_info("Termos & Privacidade", text)

    def _show_info(self, title: str, message: str):
        dialog = ctk.CTkToplevel(self)
        dialog.title(title)
        dialog.geometry("560x420")
        dialog.transient(self)
        dialog.grab_set()
        text = ctk.CTkTextbox(dialog, wrap="word", font=("Segoe UI", 12))
        text.pack(fill="both", expand=True, padx=12, pady=12)
        text.insert("1.0", message)
        text.configure(state="disabled")
        ctk.CTkButton(dialog, text="Fechar", command=dialog.destroy).pack(pady=8)

    # ------------------------------------------------------------------
    # UI HELPERS
    # ------------------------------------------------------------------
    def log_message(self, text, level="INFO"):
        getattr(self.logger, "info", lambda *_: None)("{}".format(text))
        self._ui(self._append_log_text, text)

    def _append_log_text(self, text):
        try:
            self.textbox_log.insert("end", text + "\n")
            self.textbox_log.see("end")
        except Exception:
            pass

    def update_status(self, text):
        self._ui(self._set_status_label, text)

    def _set_status_label(self, text):
        try:
            self.lbl_progress_status.configure(text=f"Status: {text}")
        except Exception:
            pass

    def _reset_control_to_idle(self):
        try:
            self.btn_start.configure(
                text="Iniciar Varredura", fg_color="#10b981", hover_color="#059669"
            )
        except Exception:
            pass

    def _set_progress(self, value):
        self._ui(self._set_progress_bar, value)

    def _set_progress_bar(self, value):
        try:
            self.progress_search.set(max(0.0, min(1.0, float(value))))
        except Exception:
            pass

    def _render_mini_image(self, jpeg_bytes: bytes):
        try:
            img = Image.open(io.BytesIO(jpeg_bytes))
            img = img.resize((420, 240))
            ctk_img = ctk.CTkImage(light_image=img, dark_image=img, size=(420, 240))
            self.lbl_mini_player.configure(image=ctk_img, text="")
        except Exception:
            pass

    def _update_ram(self, percent, used_gb, total_gb):
        try:
            if percent < 65:
                bar_color = "#34d399"
            elif percent < 85:
                bar_color = "#fbbf24"
            else:
                bar_color = "#f87171"
            self.progress_ram.configure(progress_color=bar_color)
            self.progress_ram.set(percent / 100.0)
            self.lbl_ram_status.configure(
                text=f"Uso de Memória RAM do Sistema: {percent}% "
                     f"({used_gb:.1f}GB / {total_gb:.1f}GB)"
            )
        except Exception:
            pass

    def start_ram_monitor(self):
        def monitor_loop():
            while True:
                mem = psutil.virtual_memory()
                self._ui(
                    self._update_ram,
                    mem.percent,
                    mem.used / (1024 ** 3),
                    mem.total / (1024 ** 3),
                )
                time.sleep(1.5)

        threading.Thread(target=monitor_loop, daemon=True).start()

    # ------------------------------------------------------------------
    # CONTROLE DE EXECUÇÃO
    # ------------------------------------------------------------------
    def _collect_ui_config(self):
        config = {
            "nicho": self.var_nicho.get().strip(),
            "cidades": self.var_cidades.get().strip(),
            "max_perfis_por_cidade": max(1, _safe_int(self.var_max_perfis.get(), 60)),
            "delay_entre_perfis": max(0.2, _safe_float(self.var_delay_perfil.get(), 1.2)),
            "delay_entre_cidades": max(1.0, _safe_float(self.var_delay_cidade.get(), 8.0)),
            "jitter": DEFAULT_CONFIG["jitter"],
            "coletar_site": bool(self.var_coletar_site.get()),
            "coletar_redes": bool(self.var_coletar_redes.get()),
            "modo_headless": bool(self.var_headless.get()),
            "confirmar_uso_responsavel": bool(self.var_confirmar.get()),
        }
        return config

    def toggle_scraping(self):
        if self.is_running:
            self.is_running = False
            self._stop_requested = True
            self.btn_start.configure(text="Iniciar Varredura", fg_color="#10b981", hover_color="#059669")
            self.update_status("Parada solicitada...")
            self.log_message("[SISTEMA] Solicitação de parada enviada. Aguarde a finalização da etapa atual...")
            return

        config = self._collect_ui_config()
        if not config["nicho"] or not config["cidades"]:
            self.log_message("[ALERTA] Preencha os campos de Nicho e Cidade antes de iniciar!")
            return
        if not config["confirmar_uso_responsavel"]:
            self.log_message(
                "[ALERTA] Confirme o uso responsável e a conformidade com a LGPD para iniciar."
            )
            return

        self.config.set("nicho", config["nicho"])
        self.config.set("cidades", config["cidades"])
        self.config.set("max_perfis_por_cidade", config["max_perfis_por_cidade"])
        self.config.set("delay_entre_perfis", config["delay_entre_perfis"])
        self.config.set("delay_entre_cidades", config["delay_entre_cidades"])
        self.config.set("coletar_site", config["coletar_site"])
        self.config.set("coletar_redes", config["coletar_redes"])
        self.config.set("modo_headless", config["modo_headless"])
        self.config.save()

        lista_cidades = [c.strip() for c in config["cidades"].split(",") if c.strip()]
        self.textbox_log.delete("1.0", "end")
        self.is_running = True
        self._stop_requested = False
        self.btn_start.configure(text="Parar Varredura", fg_color="#ef4444", hover_color="#dc2626")
        self.update_status("Iniciando motor de busca...")
        self.log_message(
            f"[SISTEMA] Iniciando lote de {len(lista_cidades)} cidade(s) para o nicho "
            f"'{config['nicho']}' (máx. {config['max_perfis_por_cidade']} perfis/cidade)."
        )

        threading.Thread(
            target=self.run_async_batch_scraper, args=(config, lista_cidades), daemon=True
        ).start()

    def run_async_batch_scraper(self, config, lista_cidades):
        loop = asyncio.new_event_loop()
        asyncio.set_event_loop(loop)
        try:
            loop.run_until_complete(self.batch_scraper_worker(config, lista_cidades))
        except Exception as e:
            self.log_message(f"[ERRO DE EXECUÇÃO] {str(e)}")
            self.update_status("Erro durante execução. Verifique os logs.")
        finally:
            loop.close()

    async def batch_scraper_worker(self, config, lista_cidades):
        total_cidades = len(lista_cidades)
        for idx_c, cidade in enumerate(lista_cidades):
            if not self.is_running or self._stop_requested:
                break

            self.log_message("\n" + "=" * 56)
            self.log_message(
                f"[CIDADE {idx_c + 1}/{total_cidades}] Processando: "
                f"'{config['nicho']}' em '{cidade}'"
            )
            self.log_message("=" * 56 + "\n")

            await self.scraper_worker(config, cidade)

            if idx_c + 1 < total_cidades and self.is_running and not self._stop_requested:
                pause = config["delay_entre_cidades"]
                self.log_message(
                    f"\n[PAUSA ANTI-BLOQUEIO] Aguardando {pause:.0f}s antes da próxima cidade..."
                )
                for i in range(int(pause), 0, -1):
                    if not self.is_running or self._stop_requested:
                        break
                    self.update_status(f"Pausa entre cidades: {i}s restantes...")
                    await asyncio.sleep(1)

        self.is_running = False
        self._stop_requested = False
        self._ui(self._reset_control_to_idle)
        self._set_progress(1.0)
        self.update_status("Varredura finalizada.")
        self.log_message(
            f"[SISTEMA] {datetime.now():%H:%M:%S} Lote finalizado. "
            f"Consulte os relatórios em: {self.relatorios_dir}"
        )

    # ------------------------------------------------------------------
    # SCRAPER
    # ------------------------------------------------------------------
    async def scraper_worker(self, config, cidade):
        self.update_status(f"Abrindo Chromium ({cidade})...")
        self.log_message(f"[STATUS] Abrindo navegador para {cidade}...")

        if not self.is_running or self._stop_requested:
            return

        async with async_playwright() as p:
            headless = bool(config.get("modo_headless", True))
            browser = await p.chromium.launch(headless=headless)

            context = await browser.new_context(
                viewport={"width": 1280, "height": 800},
                user_agent=(
                    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
                    "(KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"
                ),
                locale="pt-BR",
                timezone_id="America/Sao_Paulo",
            )
            page = await context.new_page()
            await self._prepare_page(page)

            search_query = f"{config['nicho']} em {cidade}"
            self.update_status(f"Buscando em {cidade}...")
            self.log_message(f"[STATUS] Buscando no Google Maps: '{search_query}'...")

            maps_url = f"https://www.google.com/maps/search/{quote_plus(search_query)}"
            await page.goto(maps_url, wait_until="domcontentloaded")
            await self._dismiss_consent(page)
            await asyncio.sleep(3)
            await self.update_mini_player(page)

            if await self._looks_like_captcha(page):
                self.log_message("[AVISO] Possível tela de verificação/CAPTCHA detectada. "
                                 "Reduza o volume e aumente as pausas.")
                await browser.close()
                return

            place_urls = await self._collect_place_urls(page)
            total = len(place_urls)
            if total == 0:
                self.log_message("[AVISO] Nenhum perfil encontrado nesta cidade.")
            else:
                self.log_message(f"[INFO] Encontrados {total} perfis em {cidade}.")

            leads = []
            max_perfis = int(config["max_perfis_por_cidade"])
            delay = float(config["delay_entre_perfis"])
            jitter = float(config.get("jitter", 0.6))

            start_time = time.time()
            processed = 0

            for idx, place_url in enumerate(place_urls):
                if not self.is_running or self._stop_requested:
                    break
                if processed >= max_perfis:
                    self.log_message(f"[INFO] Limite de {max_perfis} perfis para {cidade} atingido.")
                    break

                processed += 1
                progress = processed / max_perfis if max_perfis > 0 else 0.0
                self._set_progress(progress)

                eta_str = self._compute_eta(start_time, processed, max_perfis)
                self.update_status(
                    f"{cidade}: {processed}/{max_perfis} ({int(progress * 100)}%) | ETA: {eta_str}"
                )

                try:
                    lead = await self._extract_lead(page, place_url, config, cidade)
                    if lead:
                        leads.append(lead)
                        self.log_message(
                            f"[CAPTURADO] {lead['nome']} | Tel: {lead['telefone']} | "
                            f"Email: {lead['email']} | Insta: {lead['instagram']}"
                        )
                except Exception as exc:
                    self.log_message(f"[ERRO] Falha ao processar perfil {idx + 1}: {exc}")
                    continue

                await self.update_mini_player(page)
                jitter_sec = random.uniform(0, max(0.0, delay * jitter))
                await asyncio.sleep(delay + jitter_sec)

            await browser.close()

            # Enriquecimento simultâneo dos sites (se habilitado)
            if config.get("coletar_site", True) and leads:
                self.update_status("Enriquecendo dados dos sites...")
                self.log_message("[STATUS] Enriquecendo e-mails e redes sociais dos sites capturados...")
                await self.fetch_site_info_many(leads, config)

            for lead in leads:
                self.ram_buffer.append(lead)
            self.flush_buffer_to_db()

            if leads:
                saved_path = self.export_to_excel(config, cidade, leads)
                self.log_message(
                    f"[SISTEMA] {datetime.now():%H:%M:%S} Relatório de {cidade} salvo em: {saved_path}"
                )
            else:
                self.log_message(
                    f"[AVISO] Nenhum lead novo capturado em {cidade}. Nenhum relatório foi gerado."
                )

    def _compute_eta(self, start_time, processed, total) -> str:
        if processed <= 1:
            return "Calculando..."
        elapsed = time.time() - start_time
        avg = elapsed / processed
        eta_seconds = int((total - processed) * avg)
        return f"{eta_seconds // 60:02d}m {eta_seconds % 60:02d}s"

    async def _prepare_page(self, page):
        try:
            await page.add_init_script(
                """
                Object.defineProperty(navigator, 'webdriver', { get: () => undefined });
                Object.defineProperty(navigator, 'languages', { get: () => ['pt-BR', 'pt', 'en-US'] });
                Object.defineProperty(navigator, 'plugins', { get: () => [1, 2, 3, 4, 5] });
                window.chrome = window.chrome || {};
                window.chrome.runtime = window.chrome.runtime || {};
                """
            )
        except Exception:
            pass

    async def _dismiss_consent(self, page):
        selectors = [
            "button:has-text('Aceitar')",
            "button:has-text('Aceito')",
            "button:has-text('Concordo')",
            "button:has-text('Accept all')",
            "button[aria-label*='Aceitar']",
            "form[action*='consent'] button",
        ]
        for selector in selectors:
            try:
                el = await page.query_selector(selector)
                if el:
                    await el.click(timeout=1500)
                    await asyncio.sleep(0.5)
                    return
            except Exception:
                continue

    async def _looks_like_captcha(self, page) -> bool:
        try:
            for selector in ["#captcha-form", "form[action*='sorry']", "iframe[src*='recaptcha']"]:
                if await page.query_selector(selector):
                    return True
        except Exception:
            return False
        return False

    async def _collect_place_urls(self, page):
        selectors = ["a[href*='/maps/place/']", "a[href*='google.com/maps/place/']"]
        place_urls = []
        feed_found = False
        try:
            feed = await page.query_selector("div[role='feed']")
            if feed:
                feed_found = True
                for _ in range(5):
                    await page.eval_on_selector("div[role='feed']", "el => el.scrollBy(0, 1000)")
                    await asyncio.sleep(1.2)
                    await self.update_mini_player(page)
        except Exception:
            self.log_message("[AVISO] Painel lateral de rolagem não encontrado. "
                             "Tentando perfil único/links diretos.")

        for selector in selectors:
            elements = await page.query_selector_all(selector)
            for el in elements:
                href = await el.get_attribute("href")
                if href and href not in place_urls:
                    place_urls.append(href)

        if not place_urls:
            current_url = page.url
            if "/maps/place/" in current_url:
                self.log_message("[INFO] Resultado único detectado. Processando perfil direto...")
                place_urls.append(current_url)

        if not feed_found and not place_urls:
            self.log_message("[AVISO] Nenhum resultado visível. O Google pode ter mudado o layout.")

        return list(dict.fromkeys(place_urls))

    async def _extract_lead(self, page, place_url, config, cidade) -> dict | None:
        place_id = _normalize_url(place_url)
        if not place_id or self.is_lead_processed(place_id):
            self.log_message("[PULADO] Empresa já cadastrada no banco SQLite.")
            return None

        current = _normalize_url(page.url)
        if current != place_id:
            await page.goto(place_url, wait_until="domcontentloaded")
            await asyncio.sleep(1.8)

        nome = "N/D"
        nome_el = await page.query_selector("h1.DUwDvf, h1")
        if nome_el:
            nome = (await nome_el.inner_text()).strip()
        if _is_placeholder_name(nome):
            return None

        telefone = await self._extract_phone_from_page(page)
        site = "N/D"
        site_el = await page.query_selector("a[data-item-id='authority']")
        if site_el:
            site = await site_el.get_attribute("href") or "N/D"

        # Enriquecimento de site/redes é feito em lote (fetch_site_info_many)
        # para reduzir chamadas de rede e ser mais rápido.
        email = instagram = facebook = linkedin = "N/D"
        whatsapp_link = self._build_whatsapp_link(telefone)

        return {
            "id": place_id,
            "nome": nome,
            "telefone": telefone,
            "email": email,
            "site": site,
            "whatsapp": whatsapp_link,
            "instagram": instagram,
            "facebook": facebook,
            "linkedin": linkedin,
            "nicho": config["nicho"],
            "cidade": cidade,
            "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        }

    async def _extract_phone_from_page(self, page) -> str:
        selectors = [
            "button[data-item-id^='phone:tel:']",
            "button[data-item-id*='phone:tel:']",
            "a[href^='tel:']",
        ]
        for selector in selectors:
            try:
                el = await page.query_selector(selector)
                if not el:
                    continue
                raw = await el.get_attribute("data-item-id") or await el.get_attribute("href") or ""
                digits = _extract_phone(raw.replace("phone:tel:", "").replace("tel:", ""))
                if digits != "N/D":
                    return digits
            except Exception:
                continue
        return "N/D"

    def _build_whatsapp_link(self, telefone: str) -> str:
        if telefone == "N/D":
            return "N/D"
        digits = _extract_phone(telefone)
        if digits == "N/D":
            return "N/D"
        if len(digits) in (10, 11):
            return f"https://wa.me/55{digits}"
        return "N/D"

    async def fetch_site_info_many(self, leads, config):
        sem = asyncio.Semaphore(3)
        collect_social = bool(config.get("coletar_redes", True))

        async def enrich(item):
            if item.get("site", "N/D") in ("N/D", "") or not str(item["site"]).startswith("http"):
                return
            async with sem:
                email, instagram, facebook, linkedin = await self.fetch_site_info(
                    item["site"], collect_social=collect_social
                )
                item["email"] = email
                if collect_social:
                    item["instagram"] = instagram
                    item["facebook"] = facebook
                    item["linkedin"] = linkedin

        await asyncio.gather(*(enrich(item) for item in leads))

    async def fetch_site_info(self, url, collect_social: bool = True):
        email = instagram = facebook = linkedin = "N/D"
        try:
            async with httpx.AsyncClient(
                timeout=6.0, follow_redirects=True,
                headers={"User-Agent": "Mozilla/5.0 (compatible; LeadHunter/2.0)"},
            ) as client:
                resp = await client.get(url)
                html = resp.text

                for match in EMAIL_REGEX.findall(html or ""):
                    if not _is_noise_email(match):
                        email = match
                        break

                if collect_social:
                    soup = BeautifulSoup(html, "html.parser")
                    for a in soup.find_all("a", href=True):
                        href = a["href"].strip()
                        if instagram == "N/D" and "instagram.com" in href and "instagram.com" in href:
                            instagram = href
                        elif facebook == "N/D" and "facebook.com" in href:
                            facebook = href
                        elif linkedin == "N/D" and "linkedin.com" in href:
                            linkedin = href
        except Exception:
            pass
        return email, instagram, facebook, linkedin

    async def update_mini_player(self, page):
        try:
            screenshot_bytes = await page.screenshot(type="jpeg", quality=40)
            self._ui(self._render_mini_image, screenshot_bytes)
        except Exception:
            pass

    # ------------------------------------------------------------------
    # EXPORTAÇÃO
    # ------------------------------------------------------------------
    def export_to_excel(self, config, cidade, leads):
        data_hora = datetime.now().strftime("%Y-%m-%d_%H-%M")
        nicho_clean = config["nicho"].replace(" ", "_").replace("/", "-")
        cidade_clean = cidade.replace(" ", "_").replace("/", "-")

        pasta_dia = self.relatorios_dir / datetime.now().strftime("%Y-%m-%d")
        pasta_dia.mkdir(parents=True, exist_ok=True)
        filename = f"{nicho_clean}_{cidade_clean}_{data_hora}.xlsx"
        filepath = pasta_dia / filename

        header_fill = PatternFill("solid", fgColor="1E293B")
        header_font = Font(bold=True, color="FFFFFF")
        accent_fill = PatternFill("solid", fgColor="E2E8F0")
        center = Alignment(horizontal="center", vertical="center")

        wb = Workbook()

        # ABA 1: Visão Geral do Lead
        ws1 = wb.active
        ws1.title = "Leads Geral"
        headers1 = [
            "Timestamp", "Nome da Empresa", "Telefone", "E-mail", "Website",
            "Link WhatsApp", "Instagram", "Facebook", "LinkedIn",
            "Nicho", "Cidade",
        ]
        ws1.append(headers1)
        for lead in leads:
            ws1.append([
                lead["timestamp"], lead["nome"], lead["telefone"], lead["email"],
                lead["site"], lead["whatsapp"], lead["instagram"], lead["facebook"],
                lead["linkedin"], lead["nicho"], lead["cidade"],
            ])

        # ABA 2: Foco em Redes Sociais
        ws2 = wb.create_sheet(title="Redes Sociais")
        headers2 = [
            "Nome da Empresa", "Instagram", "Facebook", "LinkedIn",
            "Website", "Link WhatsApp", "Telefone",
        ]
        ws2.append(headers2)
        for lead in leads:
            ws2.append([
                lead["nome"], lead["instagram"], lead["facebook"], lead["linkedin"],
                lead["site"], lead["whatsapp"], lead["telefone"],
            ])

        # ABA 3: Resumo
        ws3 = wb.create_sheet(title="Resumo")
        total = len(leads)
        resumo = [
            ("Nicho", config["nicho"]),
            ("Cidade", cidade),
            ("Gerado em", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
            ("Total de leads", total),
            ("Com telefone", sum(1 for l in leads if l["telefone"] != "N/D")),
            ("Com e-mail", sum(1 for l in leads if l["email"] != "N/D")),
            ("Com WhatsApp", sum(1 for l in leads if l["whatsapp"] != "N/D")),
            ("Com Instagram", sum(1 for l in leads if l["instagram"] != "N/D")),
            ("Com Facebook", sum(1 for l in leads if l["facebook"] != "N/D")),
            ("Com LinkedIn", sum(1 for l in leads if l["linkedin"] != "N/D")),
        ]
        ws3.append(["Indicador", "Valor"])
        for row in resumo:
            ws3.append(list(row))

        for ws in (ws1, ws2, ws3):
            self._style_sheet(ws, header_fill, header_font, accent_fill, center)
        self._style_sheet(ws3, header_fill, header_font, accent_fill, center)

        wb.save(filepath)
        return str(filepath)

    def _style_sheet(self, ws, header_fill, header_font, accent_fill, center):
        try:
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = center
            ws.freeze_panes = "A2"
            ws.auto_filter.ref = ws.dimensions
            for col_idx in range(1, ws.max_column + 1):
                letter = get_column_letter(col_idx)
                max_len = 0
                for cell in ws[letter]:
                    value = cell.value
                    if value is None:
                        continue
                    max_len = max(max_len, len(str(value)))
                ws.column_dimensions[letter].width = min(max(12, max_len + 2), 45)
        except Exception:
            pass


if __name__ == "__main__":
    app = LeadHunterApp()
    app.mainloop()
