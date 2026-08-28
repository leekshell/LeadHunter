"""Testes para o App.py do LeadHunter.

Cobre as partes testáveis sem navegador/sem display:
utilitários puros, config persistente, SQLite (dedup),
exportação Excel e enriquecimento de sites (contra um
servidor HTTP local com HTML de fixture).
"""

import asyncio
import http.server
import sqlite3
import threading
from collections import deque
from pathlib import Path

import pytest
from openpyxl import load_workbook

import sys

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))

import App as lh  # noqa: E402


# ---------------------------------------------------------------------------
# fixtures
# ---------------------------------------------------------------------------
@pytest.fixture()
def stub_app(tmp_path):
    """Instância de LeadHunterApp sem passar pelo __init__ (sem GUI)."""
    app = object.__new__(lh.LeadHunterApp)
    app.base_dir = tmp_path
    app.logs_dir = tmp_path / "logs"
    app.relatorios_dir = tmp_path / "relatorios"
    app.logs_dir.mkdir()
    app.relatorios_dir.mkdir()
    app.db_path = tmp_path / "test.db"
    app.ram_buffer = deque()
    app.init_db()
    return app


def make_lead(**over):
    lead = {
        "id": "https://maps.google.com/maps/place/empresa-x",
        "nome": "Empresa X",
        "telefone": "19999998888",
        "email": "N/D",
        "site": "https://empresax.com.br",
        "whatsapp": "https://wa.me/5519999998888",
        "instagram": "N/D",
        "facebook": "N/D",
        "linkedin": "N/D",
        "nicho": "Odontologia",
        "cidade": "Campinas",
        "timestamp": "2026-08-28 10:00:00",
    }
    lead.update(over)
    return lead


FIXTURE_HTML = """
<html><head><title>Empresa X</title></head><body>
<p>Fale conosco: <a href="mailto:contato@empresax.com.br">contato@empresax.com.br</a></p>
<script>var suporte = "suporte@example.com";</script>
<a href="https://www.instagram.com/empresax.oficial/">Instagram</a>
<a href="https://www.facebook.com/empresax">Facebook</a>
<a href="https://www.linkedin.com/company/empresax">LinkedIn</a>
</body></html>
"""


@pytest.fixture(scope="module")
def fixture_server():
    """Servidor HTTP local que serve a página de fixture."""

    class Handler(http.server.BaseHTTPRequestHandler):
        def do_GET(self):
            body = FIXTURE_HTML.encode("utf-8")
            self.send_response(200)
            self.send_header("Content-Type", "text/html; charset=utf-8")
            self.send_header("Content-Length", str(len(body)))
            self.end_headers()
            self.wfile.write(body)

        def log_message(self, *args):  # silêncio
            pass

    server = http.server.ThreadingHTTPServer(("127.0.0.1", 0), Handler)
    port = server.server_address[1]
    thread = threading.Thread(target=server.serve_forever, daemon=True)
    thread.start()
    yield f"http://127.0.0.1:{port}"
    server.shutdown()


# ---------------------------------------------------------------------------
# utilitários puros
# ---------------------------------------------------------------------------
class TestPureUtils:
    def test_normalize_url_remove_query_e_fragmento(self):
        url = "https://www.google.com/maps/place/Cl%C3%ADnica/@-22.9,-47.0,17z?utm_source=x#menu"
        assert lh._normalize_url(url) == "https://www.google.com/maps/place/Cl%C3%ADnica/@-22.9,-47.0,17z"

    def test_normalize_url_vazia(self):
        assert lh._normalize_url("") == ""

    def test_extract_phone_simples(self):
        assert lh._extract_phone("(19) 3232-4455") == "1932324455"

    def test_extract_phone_remove_zero_a_esquerda(self):
        assert lh._extract_phone("019 99999-8888") == "19999998888"

    def test_extract_phone_vazio(self):
        assert lh._extract_phone("") == "N/D"
        assert lh._extract_phone("sem telefone") == "N/D"

    def test_noise_emails(self):
        assert lh._is_noise_email("suporte@example.com")
        assert lh._is_noise_email("img@logo.png")
        assert lh._is_noise_email("no-reply@site.com.br")
        assert not lh._is_noise_email("contato@empresax.com.br")

    def test_placeholder_names(self):
        assert lh._is_placeholder_name("")
        assert lh._is_placeholder_name("Google Maps")
        assert not lh._is_placeholder_name("Clínica Sorriso")

    def test_safe_int_float(self):
        assert lh._safe_int("abc", 60) == 60
        assert lh._safe_int("15", 60) == 15
        assert lh._safe_float("x,y", 1.2) == 1.2
        assert lh._safe_float("2.5", 1.2) == 2.5


# ---------------------------------------------------------------------------
# AppConfig
# ---------------------------------------------------------------------------
class TestAppConfig:
    def test_defaults_quando_nao_existe(self, tmp_path):
        cfg = lh.AppConfig(tmp_path / "config.json")
        assert cfg.get("max_perfis_por_cidade") == 60
        assert cfg.get("modo_headless") is True

    def test_roundtrip(self, tmp_path):
        path = tmp_path / "config.json"
        cfg = lh.AppConfig(path)
        cfg.set("nicho", "Odontologia")
        cfg.set("max_perfis_por_cidade", 42)
        cfg.save()
        cfg2 = lh.AppConfig(path)
        assert cfg2.get("nicho") == "Odontologia"
        assert cfg2.get("max_perfis_por_cidade") == 42

    def test_json_corrompido_nao_quebra(self, tmp_path):
        path = tmp_path / "config.json"
        path.write_text("{ json quebrado !!!", encoding="utf-8")
        cfg = lh.AppConfig(path)
        assert cfg.get("delay_entre_perfis") == 1.2

    def test_chave_desconhecida_ignorada(self, tmp_path):
        path = tmp_path / "config.json"
        path.write_text('{"chave_maluca": 1, "nicho": "Pet"}', encoding="utf-8")
        cfg = lh.AppConfig(path)
        assert cfg.get("nicho") == "Pet"
        assert cfg.get("chave_maluca") is None


# ---------------------------------------------------------------------------
# WhatsApp link (bug conhecido: telefone com DDI 55 vira N/D)
# ---------------------------------------------------------------------------
class TestWhatsAppLink:
    def test_10_digitos(self, stub_app):
        assert stub_app._build_whatsapp_link("1932324455") == "https://wa.me/551932324455"

    def test_11_digitos(self, stub_app):
        assert stub_app._build_whatsapp_link("19999998888") == "https://wa.me/5519999998888"

    def test_sem_telefone(self, stub_app):
        assert stub_app._build_whatsapp_link("N/D") == "N/D"

    def test_bug_telefone_com_ddi_55(self, stub_app):
        """Google Maps normalmente expõe o telefone COM +55
        (ex.: phone:tel:+55-19-...). _extract_phone gera 12 dígitos
        e _build_whatsapp_link só aceita 10/11 -> retorna N/D."""
        telefone_vindo_do_maps = lh._extract_phone("+55 19 99999-8888")
        assert telefone_vindo_do_maps == "5519999998888"  # 12 dígitos
        # Comportamento ATUAL (bug): gera N/D em vez do link
        assert stub_app._build_whatsapp_link(telefone_vindo_do_maps) == "N/D"


# ---------------------------------------------------------------------------
# Banco SQLite / deduplicação
# ---------------------------------------------------------------------------
class TestDatabase:
    def test_schema_criado(self, stub_app):
        conn = sqlite3.connect(stub_app.db_path)
        tables = {r[0] for r in conn.execute("SELECT name FROM sqlite_master WHERE type='table'")}
        conn.close()
        assert {"leads", "app_settings"} <= tables

    def test_flush_e_dedup(self, stub_app):
        lead = make_lead()
        assert not stub_app.is_lead_processed(lead["id"])

        stub_app.ram_buffer.append(lead)
        stub_app.flush_buffer_to_db()
        assert stub_app.is_lead_processed(lead["id"])

        # inserir de novo: INSERT OR IGNORE deve manter 1 linha só
        stub_app.ram_buffer.append(make_lead(nome="Outro nome"))
        stub_app.flush_buffer_to_db()
        conn = sqlite3.connect(stub_app.db_path)
        (count,) = conn.execute("SELECT COUNT(*) FROM leads").fetchone()
        conn.close()
        assert count == 1

    def test_ids_diferentes_para_urls_com_query(self, stub_app):
        a = make_lead(id=lh._normalize_url("https://google.com/maps/place/X?foo=1"))
        b = make_lead(id=lh._normalize_url("https://google.com/maps/place/X?foo=2"))
        assert a["id"] == b["id"]  # normalização evita duplicata por parâmetro


# ---------------------------------------------------------------------------
# Exportação Excel
# ---------------------------------------------------------------------------
class TestExcelExport:
    def test_exporta_tres_abas_com_dados(self, stub_app):
        leads = [
            make_lead(),
            make_lead(id="https://google.com/maps/place/Y", nome="Empresa Y",
                      telefone="N/D", whatsapp="N/D"),
        ]
        config = {"nicho": "Odontologia"}
        path = stub_app.export_to_excel(config, "Campinas", leads)
        assert Path(path).exists()

        wb = load_workbook(path)
        assert wb.sheetnames == ["Leads Geral", "Redes Sociais", "Resumo"]

        ws1 = wb["Leads Geral"]
        assert ws1.max_row == 3  # header + 2 leads
        assert ws1["B2"].value == "Empresa X"
        assert ws1.freeze_panes == "A2"
        assert ws1.auto_filter.ref is not None

        ws3 = wb["Resumo"]
        resumo = {row[0]: row[1] for row in ws3.iter_rows(min_row=2, values_only=True)}
        assert resumo["Total de leads"] == 2
        assert resumo["Com telefone"] == 1
        assert resumo["Com WhatsApp"] == 1

    def test_nome_arquivo_sanitizado(self, stub_app):
        path = stub_app.export_to_excel({"nicho": "Pet/Shop"}, "São Paulo", [make_lead()])
        name = Path(path).name
        assert "/" not in name
        assert name.endswith(".xlsx")


# ---------------------------------------------------------------------------
# Enriquecimento de sites (fetch_site_info) contra servidor local
# ---------------------------------------------------------------------------
class TestSiteEnrichment:
    def test_extrai_email_e_redes(self, stub_app, fixture_server):
        email, insta, fb, li = asyncio.run(
            stub_app.fetch_site_info(fixture_server, collect_social=True)
        )
        # e-mail válido (example.com do script deve ser filtrado como ruído)
        assert email == "contato@empresax.com.br"
        assert insta.startswith("https://www.instagram.com/")
        assert fb.startswith("https://www.facebook.com/")
        assert li.startswith("https://www.linkedin.com/")

    def test_sem_redes_quando_desabilitado(self, stub_app, fixture_server):
        email, insta, fb, li = asyncio.run(
            stub_app.fetch_site_info(fixture_server, collect_social=False)
        )
        assert email == "contato@empresax.com.br"
        assert insta == fb == li == "N/D"

    def test_site_fora_do_ar_nao_quebra(self, stub_app):
        email, insta, fb, li = asyncio.run(
            stub_app.fetch_site_info("http://127.0.0.1:1/", collect_social=True)
        )
        assert (email, insta, fb, li) == ("N/D", "N/D", "N/D", "N/D")

    def test_fetch_many_enriquece_em_lote(self, stub_app, fixture_server):
        leads = [
            make_lead(site=fixture_server),
            make_lead(site="N/D", id="https://google.com/maps/place/Z"),
        ]
        asyncio.run(stub_app.fetch_site_info_many(leads, {"coletar_redes": True}))
        assert leads[0]["email"] == "contato@empresax.com.br"
        assert leads[0]["instagram"].startswith("https://www.instagram.com/")
        # lead sem site permanece intocado
        assert leads[1]["email"] == "N/D"


# ---------------------------------------------------------------------------
# ETA
# ---------------------------------------------------------------------------
class TestEta:
    def test_primeiro_item_calculando(self, stub_app):
        assert stub_app._compute_eta(0, 1, 60) == "Calculando..."

    def test_eta_formato(self, stub_app):
        import time as t
        eta = stub_app._compute_eta(t.time() - 10, 10, 60)
        assert eta.endswith("s")
        assert "m" in eta
